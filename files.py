import openpyxl as op
import os
import datetime
import pandas as pd
import numpy as np
from variables import *
def create(id):
    wb=op.Workbook()
    wb.create_sheet("book") 
    ws=wb["book"]
    ws.append(("ID","Name","Date","Finished"))
    wb.create_sheet("reminder")
    ws=wb["reminder"]
    ws.append(("ID","Reminder","Date","Finished"))
    wb.save(get_excel_path(id))

def create_list(name,id):
    if os.path.exists(get_excel_path(id)):
        wb=op.load_workbook(get_excel_path(id))
    else:
        wb=op.Workbook()
    wb.create_sheet(name) 
    ws=wb[name]
    ws.append(("ID","Name","Date","Finished"))
    wb.save(get_excel_path(id))

def remove_list(list_id,id):
    wb=op.load_workbook(get_excel_path(id))
    list_name=wb.sheetnames[list_id]
    wb.remove_sheet(list_name)
    wb.save(get_excel_path(id))

def show_list(id):
    wb=op.load_workbook(get_excel_path(id))
    list_names=wb.sheetnames
    string="\n"
    for list in list_names[1:]:
        string+=str(list_names.index(list))+"-"+list+"\n"
    return string

def get_list(id):
    wb=op.load_workbook(get_excel_path(id))
    list_names=wb.sheetnames[1:]
    return list_names

def get_help():
    with open("variables.py","r") as f:
        lines=f.readlines()
    help=""
    for line in lines:
        help+=line
    return help

def add_element(list_name,info,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb[list_name]
    ws.append((str(ws.max_row),info.capitalize(),datetime.date.today(),"N"))
    wb.save(get_excel_path(id))

def show_element(list_name,id):
    df=pd.read_excel(get_excel_path(id),sheet_name=list_name)
    pos=np.where(df["Finished"]=="N")[0]
    books=""
    for i in pos:
        books+=str(df["ID"].iloc[i])+"->"+str(df["Name"].iloc[i])+"\n"
    return books

def remove_element(list_name,element_id,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb[list_name]
    element_name=""
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value==str(element_id):
            element_name=ws.cell(row=i,column=2).value
            ws.delete_rows(i)
            break
    wb.save(get_excel_path(id))
    return element_name
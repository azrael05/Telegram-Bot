import os
from files import create
import openpyxl as op
import datetime
import pandas as pd
import numpy as np
from variables import get_excel_path
## Columns "ID","Name","Date","Finished"

def add_book(book_name,id):
    if not os.path.exists(get_excel_path(id)):
        create(id)
    wb=op.load_workbook(get_excel_path(id))
    wb.active=wb["book"]
    ws=wb.active
    ws.append((str(ws.max_row),book_name[1:].capitalize(),datetime.date.today(),"N"))
    wb.save(get_excel_path(id))


def show_books(id):
    df=pd.read_excel(get_excel_path(id),sheet_name="book")
    pos=np.where(df["Finished"]=="N")[0]
    books=""
    for i in pos:
        books+=str(df["ID"].iloc[i])+"->"+str(df["Name"].iloc[i])+"\n"
    return books

def count_books(id):
    df=pd.read_excel(get_excel_path(id),sheet_name="book")
    pos=np.where(df["Finished"]=="N")[0]
    count=pos.shape[0]
    return count

def finish_book(book_id,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb["book"]
    book_name=""
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value==str(book_id):
            ws.cell(row=i,column=4).value="Y"
            book_name=ws.cell(row=i,column=2).value
            break
    wb.save(get_excel_path(id))
    return book_name

def remove_book(book_id,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb["book"]
    book_name=""
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value==str(book_id):
            book_name=ws.cell(row=i,column=2).value
            ws.delete_rows(i)
            break
    wb.save(get_excel_path(id))
    return book_name
    
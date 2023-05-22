import os
import pandas as pd
import telebot
import numpy as np
BOT_TOKEN="5911996942:AAE9rRbyK_Y9NuvG1zBkq6xVmBznsMfsLFI"
bot = telebot.TeleBot(BOT_TOKEN)
import datetime
import openpyxl as op

def create(id):
    wb=op.Workbook()
    wb.create_sheet("book")
    ws=wb["book"]
    ws.append(("ID","Name","Date","Finished"))
    wb.create_sheet("reminder")
    ws=wb["reminder"]
    ws.append(("ID","Reminder","Day","Finished"))
    wb.save(filename="{id}.xlsx".format(id=id))

def add_book(book_name,id):
    if not os.path.exists(str(id)+".xlsx"):
        create(id)
    wb=op.load_workbook("{id}.xlsx".format(id=id))
    wb.active=wb["book"]
    ws=wb.active
    ws.append((str(ws.max_row+1),book_name[1:].capitalize(),datetime.date.today(),"N"))
    wb.save(filename="{id}.xlsx".format(id=id))


def show_books(id):
    df=pd.read_excel(str(id)+".xlsx",sheet_name="book")
    pos=np.where(df["Finished"]=="N")[0]
    books=""
    for i in pos:
        books+=str(df["ID"].iloc[i])+"->"+str(df["Name"].iloc[i])+"\n"
    return books

def add_reminder(reminder,id):
    if not os.path.exists(str(id)+".xlsx"):
        create(id)
    wb=op.load_workbook("{id}.xlsx".format(id=id))
    wb.active=wb["reminder"]
    ws=wb.active
    ws.append((str(ws.max_row+1),reminder[1:].capitalize(),datetime.date.today(),"N"))
    wb.save(filename="{id}.xlsx".format(id=id))
    

def show_reminder(id):
    df=pd.read_excel(str(id)+".xlsx",sheet_name="reminder")
    pos=np.where(df["Finished"]=="N")[0]
    reminder=""
    for i in pos:
        reminder+=str(df["ID"].iloc[i])+"->"+str(df["Reminder"].iloc[i])+"\n"
    return reminder

def count_books(id):
    df=pd.read_excel(str(id)+".xlsx",sheet_name="book")
    pos=np.where(df["Finished"]=="N")[0]
    count=pos.shape[0]
    return count

def count_reminders(id):
    df=pd.read_excel(str(id)+".xlsx",sheet_name="reminder")
    pos=np.where(df["Finished"]=="N")[0]
    count=pos.shape[0]
    return count

def finish_book(book_id,id):
    wb=op.load_workbook(str(id)+".xlsx")
    ws=wb["book"]
    for i in range(ws.max_row):
        if ws.cell(row=i,column=1)==str(book_id):
            ws.cell(row=i,column=4).value="Y"


def finish_reminder(rem_id,id):
    wb=op.load_workbook(str(id)+".xlsx")
    ws=wb["reminder"]
    for i in range(ws.max_row):
        if ws.cell(row=i,column=1)==str(rem_id):
            ws.cell(row=i,column=4).value("Y")

def remove_book(book_id,id):
    df=pd.read_excel(str(id)+".xlsx")
    df.drop(np.where(df["ID"]==str(book_id))[0],axis=0)

@bot.message_handler(commands=['start', 'hello'])
def send_welcome(message):
    bot.reply_to(message, "\n /show_books -> Show all books \n /show_reminders -> List all reminders")

@bot.message_handler(commands=['show_books'])
def send_welcome(message):
    bot.reply_to(message, show_books() if show_books() else "Nothing to read")

@bot.message_handler(commands=['show_reminders'])
def send_welcome(message):
    bot.reply_to(message, show_reminder() if show_reminder() else "Nothing to reminder")


import json
import numpy as np
@bot.message_handler(func=lambda msg: True)
def confirm_order(message):
    id=message.chat.id
    text=message.text
    text=str(text).lower()
    if text.startswith(("add book","new book")):
        book_name=text.split("book")[-1]
        add_book(book_name,id)
        bot.reply_to(message,"Book added to your list {book}. \n\nNow you have {count} books in your list.".format(book=book_name[1:],count=count_books(id)))
    
    
    elif text.startswith(("add reminder","set reminder","reminder","create reminder")):
        reminder=text.split("reminder")[-1]
        add_reminder(reminder,id)
        bot.reply_to(message,"Added reminder for {reminder}. Now there are {count} reminders in your list.".format(reminder=reminder,count=count_reminders(id)))
    
    
    elif text.startswith(("show books","show book","show all books","show all book","list books","list book","all books","all book")):
        books=show_books(id)
        bot.reply_to(message,books)
    
    
    elif text.startswith(("show reminders","show reminder","list reminders","list reminder","show all reminders","show all reminder","all reminder","all reminders")):
        reminder=show_reminder(id)
        bot.reply_to(message,reminder)

    elif text.startswith(("book finished","completed book","remove book","book remove","book completed")):
        book_id=text.split(" ")[-1]
        finish_book(book_id,id)
    
    elif text.startswith("reminder remove"):
        rem_id=text.split(" ")[-1]
        finish_book(rem_id,id)
           
bot.infinity_polling()

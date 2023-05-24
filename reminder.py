import openpyxl as op
import datetime
import pandas as pd
import numpy as np
import os
from files import create
from variables import get_excel_path
## Columns "ID","Reminder","Day","Finished"


def add_reminder(reminder,id):
    if not os.path.exists(get_excel_path(id)):
        create(id)
    wb=op.load_workbook("{id}.xlsx".format(id=id))
    wb.active=wb["reminder"]
    ws=wb.active
    ws.append((str(ws.max_row),reminder[1:].capitalize(),datetime.date.today(),"N"))
    wb.save(get_excel_path(id))
    

def show_reminder(id):
    df=pd.read_excel(get_excel_path(id),sheet_name="reminder")
    pos=np.where(df["Finished"]=="N")[0]
    reminder=""
    for i in pos:
        reminder+=str(df["ID"].iloc[i])+"->"+str(df["Reminder"].iloc[i])+"\n"
    return reminder



def count_reminders(id):
    df=pd.read_excel(get_excel_path(id),sheet_name="reminder")
    pos=np.where(df["Finished"]=="N")[0]
    count=pos.shape[0]
    return count



def finish_reminder(rem_id,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb["reminder"]
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value==str(rem_id):
            rem=ws.cell(row=i,column=2).value
            ws.cell(row=i,column=4).value="Y"
    wb.save(get_excel_path(id))
    return rem


def remove_reminder(rem_id,id):
    wb=op.load_workbook(get_excel_path(id))
    ws=wb["book"]
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i,column=1).value==str(rem_id):
            rem=ws.cell(row=i,column=2).value
            ws.delete_rows(i)
            break
    wb.save(get_excel_path(id))
    return rem